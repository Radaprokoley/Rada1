import openpyxl as openpyxl
import pandas as pd

df = pd.read_excel('/Users/vidya.sasikumar/Desktop/summary files/hc_summary_prod_mapped_sorted.xls', usecols="A")
df2 = pd.read_excel('/Users/vidya.sasikumar/Desktop/summary files/hc_summary_validation_sorted.xls', usecols="A")
df = pd.DataFrame(df, columns=['Sample ID'])
df2 = pd.DataFrame(df2, columns=['Sample ID'])
df.to_excel("summary_prod_sample_list.xlsx")
df2.to_excel("summary_validation_sample_list.xlsx")
wrkbk = openpyxl.load_workbook("summary_prod_sample_list.xlsx")
sh = wrkbk.active
sample_id_list_prod = []
for i in range(2, sh.max_row + 1):
    sample_id_obj = sh.cell(row=i, column=2)
    sample_id = sample_id_obj.value
    sample_id_list_prod.append(sample_id)
sample_id_list_prod = list(set(sample_id_list_prod))
sample_id_list_prod.sort()
print("Total number of Sample ID's in prod:{}".format(len(sample_id_list_prod)))
print("summary_prod_list:{}".format(sample_id_list_prod))

wrkbk = openpyxl.load_workbook("summary_validation_sample_list.xlsx")
sh = wrkbk.active
sample_id_list_valid = []
for i in range(2, sh.max_row + 1):
    sample_id_obj = sh.cell(row=i, column=2)
    sample_id = sample_id_obj.value
    sample_id_list_valid.append(sample_id)
sample_id_list_valid = list(set(sample_id_list_valid))
sample_id_list_valid.sort()
print("Total number of Sample ID's in validation:{}".format(len(sample_id_list_valid)))
print("summary_validation_list:{}".format(sample_id_list_valid))
common_list = set(sample_id_list_prod).intersection(sample_id_list_valid)

set_difference = set(sample_id_list_valid) - set(common_list)
list_difference = list(set_difference)

print("Number of Sample ID's present in both prod and List :{}".format(len(common_list)))
print("The Sample ID's present in both Prod and Validation are:{}".format(common_list))
print("Number of Sample ID's missing from validation :{}".format(len(list_difference)))
print("The Sample ID's missing from Validation are:{}".format(list_difference))
