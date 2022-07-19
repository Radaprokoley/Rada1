import openpyxl as openpyxl
import pandas as pd

df = pd.read_excel('/Users/vidya.sasikumar/Desktop/summary files/hc_summary_prod_mapped_sorted.xls', usecols="A,C,E")
df2 = pd.read_excel('/Users/vidya.sasikumar/Desktop/summary files/hc_summary_validation_sorted.xls', usecols="A,C,E")
df = pd.DataFrame(df, columns=['Sample ID', 'Test Code', 'Marker Code'])
df2 = pd.DataFrame(df2, columns=['Sample ID', 'Test Code', 'Marker Code'])
df.to_excel("summary_prod_sample_list.xlsx")
df2.to_excel("summary_validation_sample_list.xlsx")

wrkbk = openpyxl.load_workbook("summary_prod_sample_list.xlsx")
sh = wrkbk.active
sample_id_list_prod = []

for i in range(2, sh.max_row + 1):
    sample_id_w = sh.cell(row=i, column=2).value
    sample_id_list_prod.append(sample_id_w)
sample_id_list_prod = list(set(sample_id_list_prod))
sample_id_list_prod.sort()

wrkbk = openpyxl.load_workbook("summary_validation_sample_list.xlsx")
sh = wrkbk.active
sample_id_list_valid = []
final_list_valid = []
for i in range(2, sh.max_row + 1):
    sample_id_w = sh.cell(row=i, column=2).value
    sample_id_list_valid.append(sample_id_w)
sample_id_list_valid = list(set(sample_id_list_valid))
sample_id_list_valid.sort()
common_list = set(sample_id_list_prod).intersection(sample_id_list_valid)
common_list = list(common_list)
print("Sample ID's common in both Validation and Prod are:{}".format(common_list))
print("########################################################")
wrkbk = openpyxl.load_workbook("summary_prod_sample_list.xlsx")
sh = wrkbk.active
final_list_prod = []
for id in common_list:
    marker_code_list = []
    prod_sample_marker_code = {}
    for i in range(2, sh.max_row + 1):
        sample_id = sh.cell(row=i, column=2).value
        if id == sample_id:
            marker_code = sh.cell(row=i, column=4).value
            marker_code_list.append(marker_code)
            prod_sample_marker_code[id] = marker_code_list
    final_list_prod.append(prod_sample_marker_code)
# print("The length of production mapping with codes:{}".format(len(final_list_prod)))
# print("The production mapping with codes:{}".format(final_list_prod))

wrkbk = openpyxl.load_workbook("summary_validation_sample_list.xlsx")
sh = wrkbk.active
final_list_valid = []
for id in common_list:
    marker_code_list = []
    valid_sample_marker_code = {}
    for i in range(2, sh.max_row + 1):
        sample_id = sh.cell(row=i, column=2).value
        if id == sample_id:
            marker_code = sh.cell(row=i, column=4).value
            marker_code_list.append(marker_code)
            valid_sample_marker_code[id] = marker_code_list
    final_list_valid.append(valid_sample_marker_code)
# print("The length of validation mapping with codes:{}".format(len(final_list_valid)))
# print("The validation mapping with codes:{}".format(final_list_valid))
j = 0
for k in range(len(common_list)):
    if set(final_list_prod[j][common_list[k]]).issubset(set(final_list_valid[j][common_list[k]])):
        print("{} ### The sample ID:{} in prod is a subset of Validation Marker Code set".format(j, common_list[k]))
        print("Prod set- {} ".format(final_list_prod[j][common_list[k]]))
        print("Validation set - {}".format(final_list_valid[j][common_list[k]]))
        j = j+1
    else:
        print("The sample ID:{} in prod is not a subset of Validation Marker Code set.".format([common_list[k]]))
        j = j + 1
